import tkinter as tk
import tkinter.ttk as ttk
import pygubu
import old_model

from openpyxl import load_workbook
from tkinter import PhotoImage , messagebox

from model import Model
from ttkbootstrap import Style
from datetime import date

class Formatadata(object):
    def __init__(self,  data, entrada):
        if entrada =="calendario":
            #recebe (year=2000 , month=0 , day = 1)

            day = str(data.day) if len(str(data.day))>1 else '0'+str(data.day)
            month = str(data.month+1) if len(str(data.month+1))>1 else '0'+str(data.month+1) 
            
            self.data = day+"/"+month+"/"+str(data.year)
            self.db= str(data.year)+"/"+month+"/"+day
            self.calendario = data
        
        if entrada == "interface":
            #recebe dd/mm/aaaa
            self.calendario = data

            db = data.split("/")
            db.reverse()
            self.db= "-".join(db)
            self.data = data
        
        if entrada=="db":
            self.db = data
            self.calendario = data

            data = data.split("-")
            data.reverse()
            self.data= "/".join(data)


class Busca():
    def __init__(self, master,dados,**options):
        

        self.mainwindow = master.get_object('busca_empresa_toplevel')
        
        if not master:
            master = options.get('parent')
        self.master = master
        self.options = options

        self.dados = dados
        self.madruga=Model()

        self.busca_empresa_entry=master.get_object("busca_empresa_entry")
        self.incluir_empresa_lista =  master.get_object('incluir_empresa_lista')
        self.itens_empresa =  master.get_object('itens_empresa')
        self.busca()


    def busca(self):

        busca =self.busca_empresa_entry.get()        
        dados =self.madruga.listar_empresa_filtro(busca)
        

        remove = self.incluir_empresa_lista.get_children()
        
        if remove !=():
            for x in remove:    
                self.incluir_empresa_lista.delete(x)

        for x in dados:
            self.incluir_empresa_lista.insert('', tk.END, values=x)

    def selecionar_item(self,asd):
        item =self.itens_empresa.selection()[0] 

        empresa= self.itens_empresa.item(item)['values'][0]
        
        self.dados.dados_incluir['empresa_multa']=empresa

        local = self.madruga.listar_local_empresa_filtro(empresa)
        
        incluir_local=[]
        self.dados.incluir_empresa_entrada.delete(0,"end")
        
        self.dados.incluir_empresa_entrada.insert(0,empresa)
        
        self.dados.incluir_local_entrada.select_clear()
        self.dados.incluir_local_entrada.set("")
        for x in local:
            incluir_local.append(x[0]+" | "+x[1])
        
        self.dados.incluir_local_entrada.configure(values=incluir_local)
        self.mainwindow.destroy()
        


class Busca_infracao():
    
    def __init__(self, master,dados,**options):
        

        self.mainwindow = master.get_object('busca_infracao_toplevel')
        
        if not master:
            master = options.get('parent')
        self.master = master
        self.options = options

        self.dados = dados
        self.madruga=Model()


        self.incluir_infracao_entrada=master.get_object("incluir_infracao_entrada")
        
        self.incluir_infracao_lista =  master.get_object('incluir_infracao_lista')
        self.infracao_descricao_text =  master.get_object('infracao_descricao_text')


        self.itens_infracao =  master.get_object('itens_infracao')

        self.busca()

        
    def busca(self):
        busca =self.incluir_infracao_entrada.get()        
        dados =self.madruga.listar_infracao_filtro(busca)
        

        remove = self.incluir_infracao_lista.get_children()
        
        if remove !=():
            for x in remove:    
                self.incluir_infracao_lista.delete(x)


        for x in dados:
            self.incluir_infracao_lista.insert('', tk.END, values=x)

    def selecionar_infracao(self,asd):
        

        item =self.itens_infracao.selection()[0] 
        
        codigo_infracao= self.itens_infracao.item(item)['values'][0]
        descricao_infracao=self.itens_infracao.item(item)['values'][1]
        


        self.infracao_descricao_text.delete(1.0,"end")
        self.infracao_descricao_text.insert(1.0,descricao_infracao)
        

    def selecionar_infracao_fechar(self):
        item =self.itens_infracao.selection()[0] 
        
        codigo_infracao= self.itens_infracao.item(item)['values'][0]

        self.dados.dados_incluir['infracao_multa']=codigo_infracao

        self.dados.incluir_infracao_entrada.delete(0,"end")
        
        self.dados.incluir_infracao_entrada.insert(0,codigo_infracao)


        self.mainwindow.destroy()



    def selecionar_item(self,asd):
        item =self.itens_empresa.selection()[0] 

        empresa= self.itens_empresa.item(item)['values'][0]
        
        self.dados.dados_incluir['empresa_multa']=empresa

        local = self.madruga.listar_local_empresa_filtro(empresa)
        
        incluir_local=[]
        self.dados.incluir_empresa_entrada.delete(0,"end")
        
        self.dados.incluir_empresa_entrada.insert(0,empresa)
        
        self.dados.incluir_local_entrada.select_clear()
        self.dados.incluir_local_entrada.set("")
        for x in local:
            incluir_local.append(x[0]+" | "+x[1])
        
        self.dados.incluir_local_entrada.configure(values=incluir_local)
        self.mainwindow.destroy()
        

class abil:
    def __init__(self, main , usuario):
        
        self.madruga = Model(usuario)
        self.usuario = usuario
        self.mainwindow = main.get_object('principal2')
        self.tab6 = main.get_object('tab6')
        self.incluir_termo = main.get_object('incluir_termo')

        self.tree =  main.get_object('treeview2')

        self.incluir_termo =  main.get_object("incluir_termo")

        self.incluir_local_entrada =  main.get_object('incluir_local_entrada')
        
        
        

        self.lateral_termo_item =  main.get_object('lateral_termo_item')


        self.incluir_empresa_entrada =  main.get_object('incluir_empresa_entrada')
        self.incluir_infracao_entrada =  main.get_object('incluir_infracao_entrada')
        self.incluir_termo_botao =  main.get_object('incluir_termo_botao')

        self.dados_incluir = {}



        #secçao para incluir termos

        self.lateral_busca_entrada   =  main.get_object('lateral_busca_entrada')
        self.lateral_busca_botao = main.get_object('lateral_busca_botao')
        #self.lateral_fiscal_label = main.get_object('lateral_fiscal_label')
        
        #self.lateral_fiscal_label['text'] = self.usuario[0][0]+" - "+self.usuario[0][1]

        #self.lateral_staus_sistema_label = main.get_object('lateral_staus_sistema_label')
        #self.lateral_staus_sistema_progressbar = main.get_object('lateral_staus_sistema_progressbar')
        
        




        self.incluir_data_entrada=  main.get_object('incluir_data_entrada')
        self.incluir_fiscal_entrada=  main.get_object('incluir_fiscal_entrada')
        self.incluir_hora_entrada=  main.get_object('incluir_hora_entrada')
        self.incluir_observacao_entrada= main.get_object('incluir_observacao_entrada')
        self.incluir_local_infracao_entrada=  main.get_object('incluir_local_infracao_entrada')

        fiscais = self.madruga.busca_fiscal()
        fiscal_nome = []
        for x in fiscais:
            fiscal_nome.append(x[0]+" "+x[1])

        self.incluir_fiscal_entrada['values'] = fiscal_nome


        self.lateral_visualizar_botao =  main.get_object('lateral_visualizar_botao')
        


        self.incluir_gravar_botao =  main.get_object('incluir_gravar_botao')
        self.incluir_gravar_botao.configure(style='primary.TButton')

        self.incluir_cancelar_botao =  main.get_object('incluir_cancelar_botao')

        self.incluir_cancelar_botao.configure(style='primary.Outline.TButton')
        self.lateral_visualizar_botao.configure(style='success.TButton') 

        #variaveis da aba de visualização

        self.visulalizar_num_termo_spin =  main.get_object('visulalizar_num_termo_spin')
        self.visualizar_empresa_label =  main.get_object('visualizar_empresa_label')
        self.visualizar_infracao_label =  main.get_object('visualizar_infracao_label')
        self.visualizar_data_label =  main.get_object('visualizar_data_label')
        self.visualizar_localizacao_label =  main.get_object('visualizar_localizacao_label')
        self.visualizar_peso_label =  main.get_object('visualizar_peso_label')
        self.visualizar_hora_label =  main.get_object('visualizar_hora_label')
        self.visualizar_local_ocorrencia_label =  main.get_object('visualizar_local_ocorrencia_label')
        self.visualizar_fiscal_label =  main.get_object('visualizar_fiscal_label')
        self.visualizar_observacao_label =  main.get_object('visualizar_observacao_label')

        self.visualizar_descricao_mesage =  main.get_object('visualizar_descricao_mesage')
        self.visualizar_historico_treview =  main.get_object('visualizar_historico_treview')
        self.visualizar_situacao_label =  main.get_object('visualizar_situacao_label')
        self.visualizar_acao_combobox =  main.get_object('visualizar_acao_combobox')
        self.visualizar_recomendacao_label =  main.get_object('visualizar_recomendacao_label')
        self.visualizar_infracao_checkbutton =  main.get_object('visualizar_infracao_checkbutton')
        self.visualizar_empresa_checkbutton =  main.get_object('visualizar_empresa_checkbutton')
        
        self.visualizar_empresa_checkbutton.state(['selected'])
        self.visualizar_infracao_checkbutton.state(['selected'])
        

        self.visualizar_empresa_status =  True
        self.visualizar_infracao_status =  True



        termos = self.madruga.buscar_termos()
        self.visulalizar_num_termo_spin['values']=termos        
        #self.carregar_numero_termo_incluir()

        #variaveis da aba de processo

        
        self.processo_emissao_data =  main.get_object('processo_emissao_data')
        self.processo_pos_secme =  main.get_object('processo_pos_secme')
        self.processo_oficio_saexe =  main.get_object('processo_oficio_saexe')
        self.processo_conclisao_data =  main.get_object('processo_conclisao_data')

        self.processo_emissao_data_entry =  main.get_object('processo_emissao_data_entry')
        self.processo_pos_secme_entry =  main.get_object('processo_pos_secme_entry')
        self.processo_oficio_saexe_entry =  main.get_object('processo_oficio_saexe_entry')
        self.processo_conclisao_data_entry =  main.get_object('processo_conclisao_data_entry')






        #variaveis da aba de exportar

        self.exportar_i_dia_spin =  main.get_object('exportar_i_dia_spin')
        self.exportar_i_mes_spin =  main.get_object('exportar_i_mes_spin')
        self.exportar_i_ano_spin =  main.get_object('exportar_i_ano_spin')
        self.exportar_f_dia_spin =  main.get_object('exportar_f_dia_spin')
        self.exportar_f_mes_spin =  main.get_object('exportar_f_mes_spin')
        self.exportar_f_ano_spin =  main.get_object('exportar_f_ano_spin')
        
        self.exportar_hoje_inicial_checkbox =  main.get_object('exportar_hoje_inicial_checkbox')
        self.exportar_hoje_final_checkbox =  main.get_object('exportar_hoje_final_checkbox')

        self.exportar_concluido_checkbox =  main.get_object('exportar_concluido_checkbox')
        self.exportar_pendente_checkbox =  main.get_object('exportar_pendente_checkbox')
        self.exportar_cancelado_checkbox =  main.get_object('exportar_cancelado_checkbox')
        self.exportar_recurso_checkbox =  main.get_object('exportar_recurso_checkbox')
        





        #variaveis da aba de procurar
        self.procurar_nome_entrada =  main.get_object('procurar_nome_entrada')
        self.procurar_pavilhao_combobox =  main.get_object('procurar_pavilhao_combobox')
        self.procurar_subpavilhao_combobox =  main.get_object('procurar_subpavilhao_combobox')
        self.procurar_lista_treeview =  main.get_object('procurar_lista_treeview')

        dados = self.madruga.listar_pavilhao()
        pavilhao = [x[0]for x in dados]
        self.procurar_pavilhao_combobox['values'] =[" "]+ pavilhao
        
        self.win = main
        
        self.buscar_infracao_lateral()




    def exportar_termos(self):

        data_inicial = self.exportar_i_ano_spin.get()+"-"+self.exportar_i_mes_spin.get()+"-"+self.exportar_i_dia_spin.get()
        data_final = self.exportar_f_ano_spin.get()+"-"+self.exportar_f_mes_spin.get()+"-"+self.exportar_f_dia_spin.get()


        if 'selected' in self.exportar_concluido_checkbox.state():
            status_concluido = True
        else:
            status_concluido = False

            pass
        if 'selected' in self.exportar_pendente_checkbox.state():
            status_pendente = True
        else:
            status_pendente = False
            pass
        if 'selected' in self.exportar_cancelado_checkbox.state():
            status_cancelado = True
        else:
            status_cancelado = False
            pass
        if 'selected' in self.exportar_recurso_checkbox.state():
            status_recurso = True
        else:
            status_recurso = False
            pass

        dados = self.madruga.buscar_dados_exportar(data_inicial,data_final , status_concluido, status_pendente,status_cancelado,status_recurso)
        
        file = 'relatorios/tnp.xlsx'
        workbook = load_workbook(filename=file)
        sheet = workbook.active

        p = 2
        for x in dados:
            
            data = Formatadata(x[5],'db')
            x[5]= data.data

            sheet["A"+str(p)] = x[0]
            sheet["b"+str(p)] = x[1]
            sheet["c"+str(p)] = x[2]
            sheet["d"+str(p)] = x[3]
            sheet["e"+str(p)] = x[4]
            sheet["f"+str(p)] = x[5]
            sheet["g"+str(p)] = x[6]
            sheet["H"+str(p)] = x[7]
            sheet["I"+str(p)] = x[8]
            


            p = p + 1
        #a b c d e f g


        workbook.save(filename='relatorios/relatorio.xlsx')

    def hoje_inicial(self):

        if 'selected' in self.exportar_hoje_inicial_checkbox.state():
            
            
            data = date.today()
            
            self.exportar_i_dia_spin.set(data.day)
            self.exportar_i_mes_spin.set(data.month)
            self.exportar_i_ano_spin.set(data.year)
    def hoje_final(self):

        if 'selected' in self.exportar_hoje_final_checkbox.state():
            
            
            data = date.today()
            
            self.exportar_f_dia_spin.set(data.day)
            self.exportar_f_mes_spin.set(data.month)
            self.exportar_f_ano_spin.set(data.year)

    

    def carregar_visualizar_termo(self):
        



        empresa = self.visualizar_empresa_checkbutton.state()
        if "selected" in empresa:
            self.visualizar_empresa_status =  True
        else:
            self.visualizar_empresa_status =  False


        infracao = self.visualizar_infracao_checkbutton.state()
        if "selected" in infracao:
            self.visualizar_infracao_status =  True
        else:
            self.visualizar_infracao_status =  False



        termo = self.visulalizar_num_termo_spin.get()

        dados =self.madruga.busca_termo_resumo(termo)

        dados_norma = self.madruga.busca_norma_resumo(dados[0][1])

        self.visualizar_empresa_label['text'] = "Empresa: "+dados[0][0]
        self.visualizar_infracao_label['text'] = "Infração: "+dados[0][1]
        data = Formatadata(dados[0][2], 'db')
        self.visualizar_data_label['text'] = "Data: "+data.data
        self.visualizar_localizacao_label['text'] = "Localização: "+dados[0][5]

        self.visualizar_peso_label['text'] = "Peso: "+dados_norma[0][0]

        self.visualizar_peso_label.configure(style='success.Inverse.TLabel')  

        self.visualizar_hora_label['text'] = "Hora: "+dados[0][6]
        
        self.visualizar_descricao_mesage.delete(1.0, 'end')
        self.visualizar_descricao_mesage.insert(1.0, dados_norma[0][1])
        self.visualizar_hora_label['text'] = "Hora: "+dados[0][6]
        self.visualizar_fiscal_label['text'] = "Fiscal: "+dados[0][7]
        self.visualizar_observacao_label['text'] = "Observação: "+dados[0][8]
        self.visualizar_local_ocorrencia_label['text'] = "Local: "+dados[0][9]


        self.visualizar_acao_combobox['values'] = ['Notificar',"Multar","Suspender",'Cancelado',"Recurso"]

        reincidencia = self.madruga.buscar_reincidencia(dados[0][0], dados[0][1], dados[0][2])

        historico_reincidencia = self.madruga.buscar_reincidencia_historico(dados[0][0], dados[0][1], dados[0][2], self.visualizar_empresa_status , self.visualizar_infracao_status )

        remove = self.visualizar_historico_treview.get_children()
        if remove !=():
            for x in remove:    
                self.visualizar_historico_treview.delete(x)

        for x in historico_reincidencia:
            self.visualizar_historico_treview.insert('', tk.END, values=x)


        self.visualizar_situacao_label['text'] = "Situação: "+dados[0][3]

        self.visualizar_situacao_label.configure(style='success.Inverse.TLabel')  

        if dados[0][3] == "Pendente":
            self.visualizar_situacao_label.configure(style='danger.Inverse.TLabel')  



        calculo_reincidencia = self.madruga.calcular_reincidencia(str(len(reincidencia)), dados_norma[0][0])

        print(str(len(reincidencia)))
        
        print(calculo_reincidencia)
        
        if calculo_reincidencia != []:
            self.visualizar_recomendacao_label['text'] = calculo_reincidencia[0][0] +" - UFESP: "+calculo_reincidencia[0][1]+" - Valor: "+calculo_reincidencia[0][2]
            self.visualizar_recomendacao_label.configure(style='success.Inverse.TLabel')
        else:
            if len(reincidencia) > 5:
                self.visualizar_recomendacao_label['text'] = "Suspensão 3 dias - UFESP: 0 - Valor: 0"
                self.visualizar_recomendacao_label.configure(style='success.Inverse.TLabel')
            

        if dados[0][11] != None:
            self.processo_emissao_data_entry.delete(0,"end") 
            self.processo_emissao_data_entry.insert(0,dados[0][11]) 
        else:
            self.processo_emissao_data_entry.delete(0,"end") 
            self.processo_emissao_data_entry.insert(0,"Pendente") 
            

        if dados[0][12] != None:
            self.processo_pos_secme_entry.delete(0,"end")
            self.processo_pos_secme_entry.insert(0, dados[0][12])
        
        else:

            self.processo_pos_secme_entry.delete(0,"end")
            self.processo_pos_secme_entry.insert(0,"Pendente")

        
        self.processo_oficio_saexe_entry.delete(0,'end')
        self.processo_oficio_saexe_entry.insert(0,"Ainda Pendente")

        self.processo_conclisao_data_entry.delete(0,'end')
        self.processo_conclisao_data_entry.insert(0,"Ainda Pendente")


    def processo_emissao_data_gravar(self):
        dados={}


        termo=self.visulalizar_num_termo_spin.get()
        data=self.processo_emissao_data_entry.get()
        
        dados['termo_multa'] = termo
        dados['data_recebimento_multa'] = data

        self.madruga.emissao_data_aplicar(dados)


    def processo_pos_secme_gravar(self):
        dados={}

        termo=self.visulalizar_num_termo_spin.get()
        data=self.processo_pos_secme_entry.get()
        
        dados['termo_multa'] = termo
        dados['processo_multa'] = data

        self.madruga.pos_secme_aplicar(dados)



    def atualizar_historico_multa(self):

        termo = self.visulalizar_num_termo_spin.get()

        dados =self.madruga.busca_termo_resumo(termo)

        empresa = self.visualizar_empresa_checkbutton.state()
        if "selected" in empresa:
            self.visualizar_empresa_status =  True
        else:
            self.visualizar_empresa_status =  False


        infracao = self.visualizar_infracao_checkbutton.state()
        
        if "selected" in infracao:
            self.visualizar_infracao_status =  True
        else:
            self.visualizar_infracao_status =  False


        historico_reincidencia = self.madruga.buscar_reincidencia_historico(dados[0][0], dados[0][1], dados[0][2], self.visualizar_empresa_status , self.visualizar_infracao_status )
        #print(historico_reincidencia)
        
        remove = self.visualizar_historico_treview.get_children()
        if remove !=():
            for x in remove:    
                self.visualizar_historico_treview.delete(x)

        for x in historico_reincidencia:
            self.visualizar_historico_treview.insert('', tk.END, values=x)



    def relatorio_aplicar_multa(self):
        dados={}

        dados['valor_multa'] = self.visualizar_recomendacao_label['text']
        dados['status_multa'] =self.visualizar_acao_combobox.get()

        dados['termo_multa'] = self.visulalizar_num_termo_spin.get()
        
        self.madruga.multa_aplicar(dados)
        self.carregar_visualizar_termo()
        
        pass





    def imprimir_termo(self):
        
        termo = self.visulalizar_num_termo_spin.get()
        


        dados =self.madruga.busca_termo_resumo(termo)

        dados_norma = self.madruga.busca_norma_resumo(dados[0][1])

        

        #carrega o termo de base vazio
        file = 'doc/tnp.xlsx'
        workbook = load_workbook(filename=file)
        sheet = workbook.active

        #numero termo

        termo = termo.rstrip("2023")
        termo = termo.rstrip("/")
        sheet["AB2"] = termo
        



        #select permissionario, ambulante ou carregador
        sheet["G4"] = "X"





        #select ambulante
        #sheet["S4"] = "world!"
        
        #Nome empresa
        sheet["B7"] = dados[0][0]

        #Pavilhão
        local=dados[0][5]

        if "BOX" in local:
            lugar = local.split("BOX")
            sheet["B10"] = lugar[0]
            sheet["J10"] = lugar[1]
        if "MODULO" in local:
            lugar = local.split("MODULO")
            sheet["B10"] = lugar[0]
            sheet["Q10"] = lugar[1]
        if "ATÍPICO" in local:
            lugar = local.split("-")
            sheet["B10"] = lugar[0]
            sheet["Y10"] = lugar[1]

        if "VAGA" in local:
            lugar = local.split("|")
            sheet["B10"] = lugar[0]
            sheet["Y10"] = lugar[1]

            #data = Formatadata(x[5],'db')
            #x[5]= data.data

        #Observaçao multa
        sheet["B13"] = dados[0][8]



        #peso infração m=k16 g=O16 g2=S16 crime=X16

        #pode ser usado para carregador tambem, incluir mais 'or'

        peso =dados_norma[0][0]
        if peso == "leve" or peso == "leve_ambulante":
            sheet["H16"] = "X"
        if peso == "media" or peso == "media_ambulante":
            sheet["K16"] = "X"
        if peso == "grave" or peso == "grave_ambulante":
            sheet["O16"] = "X"
        if peso == "gravissima"or peso == "gravissima_ambulante":
            sheet["S16"] = "X"
        if peso == "crime" or peso == "crime_ambulante":
            sheet["X16"] = "X"
        

        #descrição infração p1
        descricao = dados_norma[0][1]

        p1 = descricao[0:90]+" - "

        sheet["F18"] = p1
        
        #descrição infração p2
        if len(descricao)>90:
            sheet["B20"] = descricao[90:210]
       
                 
        #numero norma
        norma = dados[0][1]

        if '2.5.4' in norma:

            norma  = norma.split("-")
            sheet["R22"] = "Ng - 006 | ITEM: "+norma[0]+" SUB ITEM "+norma[1]
        else:
            sheet["G4"] = " "
            sheet["S4"] = "X"
            norma  = norma.split("-")
            sheet["R22"] = "NP OP - 035 | ITEM: "+norma[0]+" SUB ITEM "+norma[1]
        

        #consertar para formata data
        data=dados[0][2]
        data=Formatadata(data, 'db')

        sheet["M24"] = data.data+","
        
        hora= dados[0][6]
        sheet["Q24"] = hora+" Hs,"

        local_infracao=dados[0][9]
        sheet["T24"] = local_infracao

        
        #reincidencia 1 = b30 , 2 = b32 , 3 = b34 , 4 = f34 , 5 = j43 , interdição = b36


        reincidencia = self.madruga.buscar_reincidencia(dados[0][0], dados[0][1], dados[0][2])
        conta = 57
        for x in reincidencia:
            data = Formatadata(x[2], 'db')
            sheet["D"+str(conta)] = x[4] + " - " + data.data
            conta+=1

        #breakpoint()
        reincidencia = str(len(reincidencia))

        
        

        if reincidencia == "0" and peso != "gravissima":
            sheet["B30"] = "X"
        if reincidencia == "1":
            sheet["B32"] = "X"
        if reincidencia == "2":
            sheet["B34"] = "X"
        if reincidencia == "3":
            sheet["F34"] = "X"
        if reincidencia == "4":
            sheet["j34"] = "X"
        if int(reincidencia) > 4 or peso == "gravissima":
            sheet["B36"] = "X"
            sheet["w36"] = "3"


        #fiscal imprimir 
        fiscal = dados[0][7]
        fiscal = fiscal.split(" ")

        sheet["B42"] = fiscal[0]+" "+fiscal[1]



        

        punicao = dados[0][10]

        punicao = punicao.split("-")

        ufesp = punicao[1].split(":")

        if reincidencia == "1":
            sheet["H32"] = ufesp[1]
        if reincidencia == "2" or reincidencia == "3" or reincidencia == "4":
            sheet["Y34"] = ufesp[1]
        if reincidencia == "5":
            sheet["w36"] = "3"
        
        
        

        
        workbook.save(filename='doc/tnp'+termo+'-2023.xlsx')
        resposta = messagebox.showinfo("Imprimir Termo", "Arquivo gerado com sucesso!")


        


    def carregar_numero_termo_incluir(self):
        
        termo = self.madruga.busca_proximo_termo()
        proximo_termo = termo[0][0].split("/")
        proximo_termo = str(int(proximo_termo[0])+1)+"/"+proximo_termo[1]
        self.incluir_termo_botao['text']= proximo_termo



    def cadastrar_multa(self):

        self.dados_incluir["local_empresa_multa"]=self.incluir_local_entrada.get()


        data=self.incluir_data_entrada.get()

        data = Formatadata(data, 'interface')


        self.dados_incluir["data_multa"]=data.db
        #ver funcao para pegar próximo numero
        
        termo = self.madruga.busca_proximo_termo()

        proximo_termo = termo[0][0].split("/")

        proximo_termo = str(int(proximo_termo[0])+1)+"/"+proximo_termo[1]
        
        self.dados_incluir["termo_multa"]=proximo_termo

        self.incluir_data_entrada.get()

        local = self.incluir_local_infracao_entrada.get()

        if local=='':
            self.dados_incluir['local_infracao_multa'] = self.incluir_local_entrada.get()
        else:
            self.dados_incluir['local_infracao_multa'] = self.incluir_local_infracao_entrada.get()

        self.dados_incluir["fiscal_multa"]=self.incluir_fiscal_entrada.get()
        
        self.dados_incluir["hora_multa"]=self.incluir_hora_entrada.get()
        
        self.dados_incluir["observacao_multa"]=  self.incluir_observacao_entrada.get("1.0","end-1c")

        usuario = "ALEX - 50420"


        self.dados_incluir["usuario_multa"] = usuario        
        resposta = self.madruga.multa_salvar(self.dados_incluir)
        
        



        termos = self.madruga.buscar_termos()
        self.visulalizar_num_termo_spin['values']=termos        
        self.carregar_numero_termo_incluir()
        
        self.pop_multa_cadastrada(resposta)

        pass
    def pop_multa_cadastrada(self, resposta):

        if resposta == True:
            resposta = messagebox.showinfo("Cadastro Termo", "Termo cadastrado com sucesso")
            self.cadastrar_multa_apagar_dados()
        else:
            resposta = messagebox.showwarning("Cadastro Termo", "Cuidado, esta notificação já existe -" + resposta[0][0])
            
    def cadastrar_multa_apagar_dados(self):
        self.incluir_empresa_entrada.delete(0,"end")

        self.incluir_local_entrada.select_clear()
        
        self.incluir_local_entrada.set("")
        
        self.incluir_fiscal_entrada.set("")

        incluir_local=[]
        self.incluir_local_entrada.configure(values=incluir_local)
        
        self.incluir_data_entrada.delete(0,"end")

        self.incluir_hora_entrada.delete(0,"end")

        self.incluir_observacao_entrada.delete(1.0,"end")
        self.incluir_infracao_entrada.delete(0,"end")
        self.incluir_local_infracao_entrada.delete(0,"end")



    def buscar_empresa(self):
        busca =self.incluir_empresa_entrada.get()        
        dados =self.madruga.listar_empresa_filtro(busca)
        

        remove = self.incluir_empresa_lista.get_children()
        
        if remove !=():
            for x in remove:    
                self.incluir_empresa_lista.delete(x)

        for x in dados:
            self.incluir_empresa_lista.insert('', tk.END, values=x)


    
    def buscar_empresa_pop(self):
        ui = pygubu.Builder()
        ui.add_from_file("window_busca_empresa.ui")        
        win =ui.get_object("busca_empresa_toplevel")
        
        ui.connect_callbacks(Busca(ui,self))

    def buscar_infracao_pop(self):
        ui = pygubu.Builder()
        ui.add_from_file("window_busca_infracao.ui")        
        win =ui.get_object("busca_infracao_toplevel")
        
        ui.connect_callbacks(Busca_infracao(ui,self))


        

       
        
        
        

    def selecionar_item(self,asd):
        item =self.itens_empresa.selection()[0] 

        empresa= self.itens_empresa.item(item)['values'][0]
        
        self.dados_incluir['empresa_multa']=empresa


        local = self.madruga.listar_local_empresa_filtro(empresa)
        
        incluir_local=[]
        
        self.incluir_local_entrada.select_clear()
        self.incluir_local_entrada.set("")
        for x in local:
            incluir_local.append(x[0]+" | "+x[1])
        
        self.incluir_local_entrada.configure(values=incluir_local)

    def selecionar_item_lateral(self,asd):
        

        item =self.lateral_termo_item.selection()[0]

        
        self.lateral_termo= self.lateral_termo_item.item(item)['values'][0]
        

        



    def selecionar_incluir_data(self,asd):
        data = self.incluir_infracao_entrada.get()
        


    #fucões lateral



    def visualizar_termo_lateral(self):
        #abre a tab de visualizar
        self.incluir_termo.select(self.incluir_termo.tabs()[2])
        #coloca o termo selecionadop no spin de visualizar

        self.visulalizar_num_termo_spin.set(self.lateral_termo)
        
        #carrega as informações da tab
        self.carregar_visualizar_termo()

    def visualizar_processo_lateral(self):
        #abre a tab de processo
        self.incluir_termo.select(self.incluir_termo.tabs()[3])



    def buscar_infracao_lateral(self):
        
        busca = self.lateral_busca_entrada.get()
        dados = self.madruga.listar_multa(busca)
        

        remove = self.tree.get_children()
        if remove !=():
            for x in remove:    
                self.tree.delete(x)
        #print(dir(self.tree))

        for x in dados:
            
            data = Formatadata(x[2],'db')
            x[2]= data.data
            self.tree.insert('', tk.END, values=x)
        print('asd')

        old_model.pprint()
    




    def procurar_empresas_pavilhao(self,asd):
        
        busca = self.procurar_nome_entrada.get()

        pavilhao = self.procurar_pavilhao_combobox.get()

        subpavilhao = self.madruga.buscar_subpavilhao(pavilhao)
        
        lista_subpavilhao = [x[0]for x in subpavilhao]


        self.procurar_subpavilhao_combobox.select_clear()
        self.procurar_subpavilhao_combobox.set("")
        

        self.procurar_subpavilhao_combobox['values']= [" "]+lista_subpavilhao
        
        

        remove = self.procurar_lista_treeview.get_children()
        if remove !=():
            for x in remove:    
                self.procurar_lista_treeview.delete(x)


        dados = self.madruga.procurar_busca_empresa(busca,pavilhao,' ')
        for x in dados:
            self.procurar_lista_treeview.insert('', tk.END, values=x)





    def procurar_empresas_subpavilhao(self,asd):
        busca = self.procurar_nome_entrada.get()

        subpavilhao = self.procurar_subpavilhao_combobox.get()
        pavilhao = self.procurar_pavilhao_combobox.get()
        
        remove = self.procurar_lista_treeview.get_children()
        if remove !=():
            for x in remove:    
                self.procurar_lista_treeview.delete(x)


        dados = self.madruga.procurar_busca_empresa(busca,pavilhao,subpavilhao)
    
        for x in dados:
            self.procurar_lista_treeview.insert('', tk.END, values=x)
    


    def procurar_empresas_nome(self):
        busca = self.procurar_nome_entrada.get()
        subpavilhao = self.procurar_subpavilhao_combobox.get()
        pavilhao = self.procurar_pavilhao_combobox.get()
        
        dados = self.madruga.procurar_busca_empresa(busca,pavilhao,subpavilhao)
        remove = self.procurar_lista_treeview.get_children()
        if remove !=():
            for x in remove:    
                self.procurar_lista_treeview.delete(x)


        dados = self.madruga.procurar_busca_empresa(busca,pavilhao,subpavilhao)
    
        for x in dados:
            self.procurar_lista_treeview.insert('', tk.END, values=x)
    






